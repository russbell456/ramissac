from __future__ import annotations

import os
import uuid
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from openpyxl.utils import range_boundaries

class ExcelRQService:
    def __init__(self):
        self.template_path = "static/templates/plantilla_rq.xlsx"
        self.output_path = "static/generados"
        os.makedirs(self.output_path, exist_ok=True)

    def escribir_seguro(self, ws, coord, valor):
        """ Escribe en la celda raíz de un bloque combinado para evitar errores de solo lectura """
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                min_col, min_row, _, _ = range_boundaries(str(merged_range))
                ws.cell(row=min_row, column=min_col).value = valor
                return
        ws[coord].value = valor

    def generar_archivo(self, datos: dict, productos: list, ruta_firma: str):
        wb = load_workbook(self.template_path)
        ws = wb["RQ (3)"] 

        # --- ENCABEZADO (Mapeo exacto según tus capturas) ---
        # NOMBRE DEL SERVICIO: Fila 8, Columna E (5)
        self.escribir_seguro(ws, "H9", datos.get('servicio'))
        
        # FECHA DE REQUERIMIENTO: Fila 8, Columna R (18)
        self.escribir_seguro(ws, "U8", datos.get('fecha'))
        
        # SOLICITANTE: Fila 9, Columna J (10)
        self.escribir_seguro(ws, "O9", datos.get('solicitante'))

        # --- TABLA DE MATERIALES (Empieza en Fila 12) ---
        fila_inicio = 12
        for i, p in enumerate(productos):
            fila = fila_inicio + i
            
            # Nº: Columna B (2)
            ws.cell(row=fila, column=2).value = i + 1
            
            # DESCRIPCIÓN: Columna C (3)
            ws.cell(row=fila, column=4).value = p.get('nombre')
            
            # U.M.: Columna I (9)
            ws.cell(row=fila, column=9).value = p.get('unidad')
            
            # CANTIDAD: Columna J (10)
            ws.cell(row=fila, column=10).value = p.get('cantidad')

        # --- FIRMA (Área de Solicitante - Fila 33 aprox) ---
        if ruta_firma and os.path.exists(ruta_firma):
            img = PILImage.open(ruta_firma).resize((135, 65))
            img.save(ruta_firma)
            excel_img = ExcelImage(ruta_firma)
            # Colocamos la firma exactamente sobre el espacio de SOLICITANTE
            ws.add_image(excel_img, 'E48') 

        nombre_final = f"RQ_{uuid.uuid4().hex[:8]}.xlsx"
        ruta_final = os.path.join(self.output_path, nombre_final)
        wb.save(ruta_final)
        return ruta_final